"""
Сервис для экспорта данных
"""

import asyncio
import logging
import os
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from database.repositories import (
    BusOwnerRepository,
    BusRepository,
    PassengerRepository,
    ReservationRepository,
    WaitingListRepository,
)
from models.entities import Bus, BusOwner, Passenger, Reservation, WaitingListRecord

logger = logging.getLogger(__name__)


class ExportService:
    """Сервис для экспорта данных"""

    def __init__(self):
        self.bus_repository = BusRepository()
        self.reservation_repository = ReservationRepository()
        self.passenger_repository = PassengerRepository()
        self.bus_owner_repository = BusOwnerRepository()
        self.waiting_list_repository = WaitingListRepository()

    async def export_buses_to_excel(self) -> str:
        """
        Экспортирует данные об автобусах в Excel файл

        Returns:
            str: Путь к временному файлу
        """
        try:
            # Получаем данные из базы
            buses = self.bus_repository.get_all()
            reservations = self.reservation_repository.get_all()
            passengers = self.passenger_repository.get_all()
            bus_owners = self.bus_owner_repository.get_all()

            # Создаем временный файл
            temp_file = f"temp_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            # Генерируем Excel файл в отдельном потоке
            await asyncio.to_thread(
                self._generate_excel_file,
                buses,
                reservations,
                passengers,
                bus_owners,
                temp_file,
            )

            return temp_file

        except Exception as e:
            logger.error(f"Ошибка при экспорте данных: {str(e)}")
            raise

    async def export_personal_data_to_excel(
        self, bus_ids: List[int], chief_view: bool = False
    ) -> str:
        """
        Экспортирует персональные данные пассажиров и лист ожидания по выбранным автобусам

        Returns:
            str: Путь к временному файлу
        """
        try:
            buses = [
                bus
                for bus in self.bus_repository.get_all()
                if bus.id in {int(bus_id) for bus_id in bus_ids}
            ]
            passengers = self.passenger_repository.get_all()
            reservations = self.reservation_repository.get_all()
            waiting_records = self.waiting_list_repository.get_waiting_records()

            temp_file = f"temp_personal_data_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            await asyncio.to_thread(
                self._generate_personal_data_excel_file,
                buses,
                passengers,
                reservations,
                waiting_records,
                temp_file,
                chief_view,
            )

            return temp_file

        except Exception as e:
            logger.error(f"Ошибка при экспорте персональных данных: {str(e)}")
            raise

    def _generate_excel_file(
        self,
        buses: List[Bus],
        reservations: List[Reservation],
        passengers: List[Passenger],
        bus_owners: List[BusOwner],
        filename: str,
    ):
        """Генерирует Excel файл с данными"""
        wb = openpyxl.Workbook()

        # Если нет автобусов, оставляем дефолтный лист с сообщением
        if not buses:
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Данные об автобусах отсутствуют"])
            wb.save(filename)
            return

        wb.remove(wb.active)  # Удаляем дефолтный лист только если есть данные

        # Группируем пассажиров по автобусам
        bus_passengers = {}
        for res in reservations:
            bus_id = str(res.bus_id)
            passenger = next(
                (p for p in passengers if str(p.id) == str(res.passenger_id)), None
            )
            if passenger:
                if bus_id not in bus_passengers:
                    bus_passengers[bus_id] = []
                bus_passengers[bus_id].append(passenger)

        # Создаем листы для каждого автобуса
        for bus in buses:
            bus_id = str(bus.id)
            sheet_name = f"Автобус {bus.number}"[:31]  # Ограничиваем длину имени листа
            ws = wb.create_sheet(title=sheet_name)

            # Добавляем заголовок с информацией об автобусе
            bus_info = [
                f"Автобус: {bus.number}",
                f"Маршрут: {bus.departure_place} - {bus.destination}",
                f"Направление: {bus.direction}",
                f"Дата/время: {bus.departure_date} {bus.departure_time}",
                f"Вместимость: {bus.capacity}",
            ]

            # Находим ответственных за автобус
            owners = [o for o in bus_owners if str(o.bus_id) == bus_id]
            responsible_persons = []
            for owner in owners:
                chief = next(
                    (p for p in passengers if str(p.id) == str(owner.chief_id)), None
                )
                if chief:
                    responsible_persons.append(
                        f"{chief.fio} (@{chief.telegram_username})"
                    )

            if responsible_persons:
                bus_info.append(f"Ответственные: {', '.join(responsible_persons)}")

            # Записываем информацию об автобусе
            for line in bus_info:
                ws.append([line])

            # Пустая строка для разделения
            ws.append([])

            # Заголовки таблицы пассажиров
            ws.append(["№", "ФИО", "Username", "Телефон", "Комментарий"])

            # Данные пассажиров
            for idx, passenger in enumerate(bus_passengers.get(bus_id, []), start=1):
                ws.append(
                    [
                        idx,
                        passenger.fio or "",
                        (
                            f"@{passenger.telegram_username}"
                            if passenger.telegram_username
                            else ""
                        ),
                        passenger.phone or "",
                        passenger.comment or "",
                    ]
                )

        wb.save(filename)

    def _generate_personal_data_excel_file(
        self,
        buses: List[Bus],
        passengers: List[Passenger],
        reservations: List[Reservation],
        waiting_records: List[WaitingListRecord],
        filename: str,
        chief_view: bool = False,
    ):
        """Генерирует Excel файл с персональными данными по выбранным автобусам"""
        wb = openpyxl.Workbook()

        if not buses:
            ws = wb.active
            ws.title = "Нет данных"
            ws.append(["Нет автобусов для выгрузки персональных данных"])
            wb.save(filename)
            return

        wb.remove(wb.active)
        passengers_by_id: Dict[int, Passenger] = {
            passenger.id: passenger
            for passenger in passengers
            if passenger.id is not None
        }

        reservations_by_bus: Dict[int, List[Reservation]] = {}
        for reservation in reservations:
            reservations_by_bus.setdefault(reservation.bus_id, []).append(reservation)

        waiting_by_bus: Dict[int, List[WaitingListRecord]] = {}
        for waiting_record in waiting_records:
            waiting_by_bus.setdefault(waiting_record.bus_id, []).append(waiting_record)

        for bus in buses:
            sheet_name = f"Автобус {bus.number}, {bus.direction}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            if chief_view:
                self._setup_chief_export_sheet(ws, bus)
                self._fill_chief_export_sheet(
                    ws,
                    bus,
                    passengers_by_id,
                    reservations_by_bus.get(bus.id, []),
                    waiting_by_bus.get(bus.id, []),
                )
                continue

            self._setup_personal_data_sheet(ws, bus)

            row_number = 1
            current_row = 6

            bus_reservations = sorted(
                reservations_by_bus.get(bus.id, []),
                key=lambda item: item.reservation_date or "",
            )
            for reservation in bus_reservations:
                passenger = passengers_by_id.get(reservation.passenger_id)
                if not passenger:
                    continue

                self._append_personal_data_row(ws, current_row, row_number, passenger)
                row_number += 1
                current_row += 1

            bus_waiting_records = sorted(
                waiting_by_bus.get(bus.id, []),
                key=lambda item: item.request_time or "",
            )
            for waiting_record in bus_waiting_records:
                passenger = passengers_by_id.get(waiting_record.passenger_id)
                if not passenger:
                    continue

                self._append_personal_data_row(ws, current_row, row_number, passenger)
                row_number += 1
                current_row += 1

            total_exported_rows = row_number - 1
            empty_rows_count = 0
            for _ in range(empty_rows_count):
                self._append_personal_data_blank_row(ws, current_row)
                current_row += 1

            if total_exported_rows == 0:
                ws.cell(
                    row=current_row,
                    column=1,
                    value="Нет пассажиров и записей в очереди",
                )
                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=9,
                )
                self._style_personal_data_empty_row(ws, current_row)

        wb.save(filename)

    def _setup_chief_export_sheet(self, ws, bus: Bus):
        """Оформляет лист выгрузки для шефа автобуса"""
        ws.cell(
            row=1,
            column=1,
            value=f"Автобус {bus.number or ''}, {bus.direction or ''}".strip(", "),
        )
        year = datetime.now().strftime("%y")
        ws.cell(
            row=2,
            column=1,
            value=f"Дата: {(bus.departure_date or '').strip()}.{year}".strip("."),
        )

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22

        ws["A1"].font = Font(name="Calibri", size=14, bold=True)
        ws["A2"].font = Font(name="Calibri", size=12, bold=True)

    def _fill_chief_export_sheet(
        self,
        ws,
        bus: Bus,
        passengers_by_id: Dict[int, Passenger],
        reservations: List[Reservation],
        waiting_records: List[WaitingListRecord],
    ):
        """Заполняет лист выгрузки для шефа: брони и лист ожидания"""
        current_row = 4

        booked_passengers = self._sort_passengers_for_chief_export(
            [
                passenger
                for reservation in reservations
                if (passenger := passengers_by_id.get(reservation.passenger_id))
            ]
        )
        current_row = self._append_chief_export_section(
            ws, current_row, "Забронированы", booked_passengers
        )

        current_row += 1

        waiting_passengers = self._sort_passengers_for_chief_export(
            [
                passenger
                for waiting_record in waiting_records
                if waiting_record.is_waiting()
                and (passenger := passengers_by_id.get(waiting_record.passenger_id))
            ]
        )
        self._append_chief_export_section(
            ws, current_row, "Лист ожидания", waiting_passengers
        )

    def _append_chief_export_section(
        self,
        ws,
        start_row: int,
        title: str,
        passengers: List[Passenger],
    ) -> int:
        """Добавляет секцию шефской выгрузки и возвращает следующую строку"""
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(name="Calibri", size=12, bold=True)

        header_row = start_row + 1
        headers = ("Фамилия", "Имя", "Отчество")
        for column, value in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=column, value=value)
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
            cell.border = self._get_personal_data_border()

        data_row = header_row + 1
        if not passengers:
            ws.cell(row=data_row, column=1, value="Нет записей")
            ws.cell(row=data_row, column=1).font = Font(
                name="Calibri", size=11, italic=True
            )
            return data_row + 1

        for passenger in passengers:
            last_name, first_name, patronymic = self._get_passenger_name_parts(
                passenger
            )
            values = (last_name, first_name, patronymic)
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=column, value=value)
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(vertical="center")
                cell.border = self._get_personal_data_border()
            data_row += 1

        return data_row

    def _sort_passengers_for_chief_export(
        self, passengers: List[Passenger]
    ) -> List[Passenger]:
        """Сортирует пассажиров по алфавиту, начиная с фамилии"""
        return sorted(
            passengers,
            key=lambda passenger: tuple(
                part.strip().lower()
                for part in self._get_passenger_name_parts(passenger)
            ),
        )

    def _setup_personal_data_sheet(self, ws, bus: Bus):
        """Оформляет лист персональной выгрузки по шаблону перевозчика"""
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=4)

        ws.cell(
            row=2,
            column=2,
            value=f"Автобус {bus.number or ''}, {bus.direction or ''}".strip(),
        )
        year = datetime.now().strftime("%y")
        ws.cell(
            row=3,
            column=2,
            value=f"Дата: {(bus.departure_date or '').strip()}.{year}".strip("."),
        )

        headers = {
            1: "№",
            2: "ФИО (полностью)",
            5: "Дата\nрождения\nчч.мм.гггг",
            6: "Серия+номер\nдокумента",
            7: "Вид документа\n(СвРжд/ПасРФ)",
            8: "Гражданство\n(если не РФ)",
            9: "контактный телефон\n(ответственного лица\nребенка/сопровождающего)",
        }
        for column, title in headers.items():
            ws.cell(row=5, column=column, value=title)

        column_widths = {
            "A": 5,
            "B": 20,
            "C": 20,
            "D": 20,
            "E": 18,
            "F": 25,
            "G": 26,
            "H": 20,
            "I": 42,
        }
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        ws.row_dimensions[2].height = 26
        ws.row_dimensions[3].height = 26
        ws.row_dimensions[5].height = 90
        ws.freeze_panes = "A6"

        title_font = Font(name="Calibri", size=16, bold=True)
        header_font = Font(name="Calibri", size=16, bold=True)
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        table_border = self._get_personal_data_border()

        for row in (2, 3):
            for cell in ws[row][1:7]:
                cell.border = table_border
            ws.cell(row=row, column=2).font = title_font
            ws.cell(row=row, column=2).alignment = center_alignment

        for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=9):
            for cell in row:
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = table_border
                cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    def _append_personal_data_row(
        self, ws, row: int, number: int, passenger: Passenger
    ):
        """Добавляет строку пассажира в форму персональной выгрузки"""
        last_name, first_name, patronymic = self._get_passenger_name_parts(passenger)
        citizenship = self._format_citizenship_for_export(passenger.citizenship)
        values = [
            number,
            last_name,
            first_name,
            patronymic,
            passenger.birth_date or "",
            passenger.passport_number or "",
            "Паспорт" if passenger.passport_number else "",
            citizenship,
            passenger.phone or "",
        ]

        for column, value in enumerate(values, start=1):
            ws.cell(row=row, column=column, value=value)

        ws.row_dimensions[row].height = 24
        table_border = self._get_personal_data_border()
        for cells in ws.iter_rows(min_row=row, max_row=row, min_col=1, max_col=9):
            for cell in cells:
                cell.border = table_border
                cell.font = Font(name="Calibri", size=14, bold=cell.column in (2, 3, 4))
                cell.alignment = Alignment(
                    horizontal="center" if cell.column in (1, 5, 6, 7, 8) else "left",
                    vertical="center",
                    wrap_text=True,
                )

    def _style_personal_data_empty_row(self, ws, row: int):
        """Оформляет строку с сообщением об отсутствии данных"""
        table_border = self._get_personal_data_border()
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name="Calibri", size=14, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for cells in ws.iter_rows(min_row=row, max_row=row, min_col=1, max_col=9):
            for item in cells:
                item.border = table_border

    def _append_personal_data_blank_row(self, ws, row: int):
        """Добавляет пустую строку в форму выгрузки"""
        ws.row_dimensions[row].height = 24
        table_border = self._get_personal_data_border()
        for column in range(1, 10):
            cell = ws.cell(row=row, column=column, value="")
            cell.border = table_border
            cell.font = Font(name="Calibri", size=14)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _get_passenger_name_parts(self, passenger: Passenger) -> tuple:
        """Возвращает фамилию, имя и отчество с fallback на поле FIO"""
        last_name = passenger.last_name or ""
        first_name = passenger.first_name or ""
        patronymic = passenger.patronymic or ""

        if not any([last_name, first_name, patronymic]) and passenger.fio:
            parts = passenger.fio.split(maxsplit=2)
            last_name = parts[0] if len(parts) > 0 else ""
            first_name = parts[1] if len(parts) > 1 else ""
            patronymic = parts[2] if len(parts) > 2 else ""

        return last_name, first_name, patronymic

    @staticmethod
    def _format_citizenship_for_export(citizenship: str) -> str:
        """Колонка гражданства заполняется только для не-РФ значений"""
        if not citizenship:
            return ""

        normalized = citizenship.strip().lower().replace(".", "")
        if normalized in {"рф", "россия", "российская федерация"}:
            return ""

        return citizenship

    @staticmethod
    def _get_personal_data_border() -> Border:
        """Возвращает границу для формы персональной выгрузки"""
        side = Side(style="thin", color="000000")
        return Border(left=side, right=side, top=side, bottom=side)

    def cleanup_temp_file(self, filepath: str):
        """Удаляет временный файл"""
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception as e:
            logger.error(f"Ошибка при удалении временного файла {filepath}: {str(e)}")
