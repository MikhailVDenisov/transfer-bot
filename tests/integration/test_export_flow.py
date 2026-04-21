"""
Интеграционные тесты для экспорта данных
"""

import os
import tempfile

import pytest

from database.repositories import BusOwnerRepository
from services.bus_service import BusService
from services.export_service import ExportService
from services.passenger_service import PassengerService
from tests.factories import BusFactory, PassengerFactory


@pytest.mark.integration
class TestExportFlow:
    """Интеграционные тесты для экспорта данных"""

    def test_export_service_with_real_data(self, temp_db):
        """Тест экспорта с реальными данными"""
        # Создаем сервисы
        export_service = ExportService()
        passenger_service = PassengerService()
        bus_service = BusService()

        # 1. Создаем пассажиров
        passengers_data = [
            (
                "user1",
                "123456789",
                "Иванов Иван Иванович",
                "+7900123456",
                "Комментарий 1",
                "user",
            ),
            (
                "user2",
                "123456790",
                "Петров Петр Петрович",
                "+7900123457",
                "Комментарий 2",
                "user",
            ),
            (
                "admin1",
                "123456791",
                "Сидоров Сидор Сидорович",
                "+7900123458",
                "Админ",
                "admin",
            ),
        ]

        for username, chat_id, fio, phone, comment, role in passengers_data:
            passenger, created = passenger_service.get_or_create_passenger(
                username, chat_id
            )
            assert created is True

            # Обновляем дополнительные данные
            from database.connection import db_connection

            db_connection.execute_query(
                "UPDATE Passengers SET FIO = ?, Phone = ?, Comment = ?, Role = ? WHERE Telegram_username = ?",
                (fio, phone, comment, role, username),
            )

        # 2. Создаем автобусы
        buses_data = [
            (
                "БУС-001",
                "Москва",
                "Переславль-Залесский",
                "2024-01-15",
                "10:00",
                30,
                "Туда",
                True,
            ),
            (
                "БУС-002",
                "Москва",
                "Переславль-Залесский",
                "2024-01-15",
                "12:00",
                25,
                "Обратно",
                True,
            ),
        ]

        for bus_data in buses_data:
            from database.connection import db_connection

            db_connection.execute_query(
                "INSERT INTO Buses (Number, Departure_Place, Destination, DepartureDate, DepartureTime, Capacity, Direction, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                bus_data,
            )

        buses = bus_service.get_all_buses()
        assert len(buses) == 2

        # 3. Создаем бронирования
        from database.connection import db_connection

        passengers = passenger_service.repository.get_all()

        # Бронируем места
        db_connection.execute_query(
            "INSERT INTO Reservations (PassengerID, BusID, ReservationDate, Direction) VALUES (?, ?, ?, ?)",
            (passengers[0].id, buses[0].id, "2024-01-15 10:00:00", buses[0].direction),
        )
        db_connection.execute_query(
            "INSERT INTO Reservations (PassengerID, BusID, ReservationDate, Direction) VALUES (?, ?, ?, ?)",
            (passengers[1].id, buses[0].id, "2024-01-15 10:00:00", buses[0].direction),
        )
        db_connection.execute_query(
            "INSERT INTO Reservations (PassengerID, BusID, ReservationDate, Direction) VALUES (?, ?, ?, ?)",
            (passengers[2].id, buses[1].id, "2024-01-15 10:00:00", buses[1].direction),
        )

        # 4. Создаем владельцев автобусов
        db_connection.execute_query(
            "INSERT INTO BusOwners (BusID, ChiefID) VALUES (?, ?)",
            (buses[0].id, passengers[2].id),  # Админ владеет первым автобусом
        )

        # 5. Экспортируем данные
        import asyncio

        temp_file = asyncio.run(export_service.export_buses_to_excel())

        # 6. Проверяем, что файл создан
        assert os.path.exists(temp_file)
        assert temp_file.endswith(".xlsx")

        # 7. Проверяем содержимое файла
        import openpyxl

        wb = openpyxl.load_workbook(temp_file)

        # Должно быть 2 листа (по одному на автобус)
        assert len(wb.sheetnames) == 2
        assert "Автобус БУС-001" in wb.sheetnames
        assert "Автобус БУС-002" in wb.sheetnames

        # Проверяем первый лист
        ws1 = wb["Автобус БУС-001"]
        rows = list(ws1.iter_rows(values_only=True))

        # Должна быть информация об автобусе
        assert any("Автобус: БУС-001" in str(row) for row in rows)
        assert any("Маршрут: Москва - Переславль-Залесский" in str(row) for row in rows)
        assert any("Направление: Туда" in str(row) for row in rows)

        # Должны быть заголовки таблицы пассажиров
        assert ("№", "ФИО", "Username", "Телефон", "Комментарий") in rows

        # Должны быть данные пассажиров
        passenger_rows = [
            row for row in rows if len(row) == 5 and isinstance(row[0], int)
        ]
        assert len(passenger_rows) == 2  # Два пассажира в первом автобусе

        # Проверяем второй лист
        ws2 = wb["Автобус БУС-002"]
        rows2 = list(ws2.iter_rows(values_only=True))

        passenger_rows2 = [
            row for row in rows2 if len(row) == 5 and isinstance(row[0], int)
        ]
        assert len(passenger_rows2) == 1  # Один пассажир во втором автобусе

        # 8. Очищаем временный файл
        export_service.cleanup_temp_file(temp_file)
        assert not os.path.exists(temp_file)

    def test_export_service_empty_data(self, temp_db):
        """Тест экспорта с пустыми данными"""
        export_service = ExportService()

        # Экспортируем данные (база пустая)
        import asyncio

        temp_file = asyncio.run(export_service.export_buses_to_excel())

        # Проверяем, что файл создан
        assert os.path.exists(temp_file)

        # Проверяем содержимое файла
        import openpyxl

        wb = openpyxl.load_workbook(temp_file)

        # Должен быть только один лист с сообщением об отсутствии данных
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Нет данных"

        # Очищаем временный файл
        export_service.cleanup_temp_file(temp_file)

    def test_export_personal_data_with_waiting_list(self, temp_db):
        """Тест выгрузки персональных данных с пассажирами и очередью"""
        export_service = ExportService()

        from database.connection import db_connection

        db_connection.execute_query(
            "INSERT INTO Buses (Number, Departure_Place, Destination, DepartureDate, DepartureTime, Capacity, Direction, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                "БУС-101",
                "Москва",
                "Переславль-Залесский",
                "2024-01-15",
                "10:00",
                30,
                "Туда",
                True,
            ),
        )

        db_connection.execute_query(
            "INSERT INTO Passengers (Telegram_username, ChatID, Role, LastName, FirstName, Patronymic, Phone, BirthDate, PassportNumber, Citizenship, PersonalDataConfirmed) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                "user_booked",
                "123",
                "user",
                "Иванов",
                "Иван",
                "Иванович",
                "+79001234567",
                "01.01.1990",
                "1234 567890",
                "РФ",
                True,
            ),
        )
        db_connection.execute_query(
            "INSERT INTO Passengers (Telegram_username, ChatID, Role, LastName, FirstName, Patronymic, Phone, BirthDate, PassportNumber, Citizenship, PersonalDataConfirmed) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                "user_waiting",
                "124",
                "user",
                "Петров",
                "Петр",
                "Петрович",
                "+79007654321",
                "02.02.1992",
                "4321 098765",
                "РФ",
                True,
            ),
        )

        bus = export_service.bus_repository.get_all()[0]
        passengers = export_service.passenger_repository.get_all()
        booked_passenger = next(
            p for p in passengers if p.telegram_username == "user_booked"
        )
        waiting_passenger = next(
            p for p in passengers if p.telegram_username == "user_waiting"
        )

        db_connection.execute_query(
            "INSERT INTO Reservations (PassengerID, BusID, ReservationDate, Direction) VALUES (?, ?, ?, ?)",
            (booked_passenger.id, bus.id, "2024-01-10 09:00:00", bus.direction),
        )
        db_connection.execute_query(
            "INSERT INTO WaitingList (PassengerID, BusID, RequestTime, Status, NotificationSent) VALUES (?, ?, ?, ?, ?)",
            (waiting_passenger.id, bus.id, "2024-01-10 10:00:00", "Waiting", "No"),
        )

        import asyncio

        import openpyxl

        temp_file = asyncio.run(export_service.export_personal_data_to_excel([bus.id]))

        assert os.path.exists(temp_file)

        wb = openpyxl.load_workbook(temp_file)
        assert "Автобус БУС-101" in wb.sheetnames

        ws = wb["Автобус БУС-101"]
        rows = list(ws.iter_rows(values_only=True))

        assert (
            "№",
            "Статус",
            "Фамилия",
            "Имя",
            "Отчество",
            "Телефон",
            "Дата рождения",
            "Паспорт",
            "Гражданство",
            "Telegram",
            "Дата добавления",
        ) in rows
        assert any(
            row[1] == "Пассажир" and row[2] == "Иванов" for row in rows if row[0]
        )
        assert any(row[1] == "Очередь" and row[2] == "Петров" for row in rows if row[0])

        export_service.cleanup_temp_file(temp_file)

    def test_export_service_large_dataset(self, temp_db):
        """Тест экспорта с большим объемом данных"""
        export_service = ExportService()

        # Создаем много пассажиров и автобусов
        from database.connection import db_connection

        # Создаем 10 автобусов
        for i in range(10):
            db_connection.execute_query(
                "INSERT INTO Buses (Number, Departure_Place, Destination, DepartureDate, DepartureTime, Capacity, Direction, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    f"БУС-{i:03d}",
                    "Москва",
                    "Переславль-Залесский",
                    "2024-01-15",
                    f"{10+i}:00",
                    30,
                    "Туда",
                    True,
                ),
            )

        # Создаем 50 пассажиров
        for i in range(50):
            db_connection.execute_query(
                "INSERT INTO Passengers (Telegram_username, ChatID, FIO, Phone, Comment, Role) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    f"user{i}",
                    f"12345678{i}",
                    f"Пользователь {i}",
                    f"+790012345{i}",
                    f"Комментарий {i}",
                    "user",
                ),
            )

        # Создаем бронирования (по 3 пассажира на автобус)
        buses = export_service.bus_repository.get_all()
        passengers = export_service.passenger_repository.get_all()

        for i, bus in enumerate(buses):
            for j in range(3):
                passenger_idx = (i * 3 + j) % len(passengers)
                db_connection.execute_query(
                    "INSERT INTO Reservations (PassengerID, BusID, ReservationDate, Direction) VALUES (?, ?, ?, ?)",
                    (
                        passengers[passenger_idx].id,
                        bus.id,
                        "2024-01-15 10:00:00",
                        bus.direction,
                    ),
                )

        # Экспортируем данные
        import asyncio

        temp_file = asyncio.run(export_service.export_buses_to_excel())

        # Проверяем, что файл создан
        assert os.path.exists(temp_file)

        # Проверяем содержимое файла
        import openpyxl

        wb = openpyxl.load_workbook(temp_file)

        # Должно быть 10 листов (по одному на автобус)
        assert len(wb.sheetnames) == 10

        # Проверяем один из листов
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

        # Должны быть данные пассажиров
        passenger_rows = [
            row for row in rows if len(row) == 5 and isinstance(row[0], int)
        ]
        assert len(passenger_rows) == 3  # По 3 пассажира на автобус

        # Очищаем временный файл
        export_service.cleanup_temp_file(temp_file)
