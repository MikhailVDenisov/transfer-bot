import os
import json
import asyncio

import gspread
import logging as logger
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputMediaPhoto
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
)
from dotenv import load_dotenv

load_dotenv(".env")
logging = logger.getLogger(__name__)

# Получение переменных из env
TOKEN = os.getenv("TELEGRAM_TOKEN")
spreadsheet_id = os.getenv("GOOGLE_OAUTH_TOKEN")
credentials_json_str = os.getenv("GOOGLE_CREDENTIALS_JSON", "")

credentials_info = json.loads(credentials_json_str)
credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_info)
client = gspread.authorize(credentials)

spreadsheet = client.open_by_key(spreadsheet_id)

# Получение листов
passengers_sheet = spreadsheet.worksheet("Passengers")
buses_sheet = spreadsheet.worksheet("Buses")
reservations_sheet = spreadsheet.worksheet("Reservations")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # Получение chat_id и username пользователя
    chat_id = None
    username = None
    if update.message:
        chat_id = update.message.chat_id
        username = update.message.from_user.username
    elif update.callback_query:
        chat_id = update.callback_query.message.chat_id
        username = update.callback_query.from_user.username

    if username:
        try:
            # Проверка, есть ли пользователь в таблице по username
            cell = passengers_sheet.find(username)
            # Если есть, обновляем ChatID, если он пустой
            headers = passengers_sheet.row_values(1)
            if "ChatID" in headers:
                chatid_col_idx = headers.index("ChatID") + 1
                current_value = passengers_sheet.cell(cell.row, chatid_col_idx).value
                if not current_value and chat_id:
                    passengers_sheet.update_cell(cell.row, chatid_col_idx, str(chat_id))
            
            # Проверяем роль пользователя
            is_admin = False
            if "Role" in headers:
                role_col_idx = headers.index("Role") + 1
                role = passengers_sheet.cell(cell.row, role_col_idx).value
                is_admin = role and role.lower() == "admin"
        except:
            # Пользователь не найден, создаем новую запись
            headers = passengers_sheet.row_values(1)
            # Получаем список существующих ID
            if "ID" in headers:
                id_col_idx = headers.index("ID") + 1
                all_ids = passengers_sheet.col_values(id_col_idx)[1:]  # пропускаем заголовок
                max_id = 0
                for val in all_ids:
                    try:
                        num = int(val)
                        if num > max_id:
                            max_id = num
                    except:
                        continue
                new_id = max_id + 1
                # Создание новой строки
                new_row = []
                for header in headers:
                    if header == "Telegram_username":
                        new_row.append(username or "")
                    elif header == "ChatID":
                        new_row.append(str(chat_id) if chat_id else "")
                    elif header == "ID":
                        new_row.append(str(new_id))
                    else:
                        new_row.append("")
                passengers_sheet.append_row(new_row)
            is_admin = False

    # Остальной код стартового сообщения
    welcome_message = (
        "Привет! Я трансфер-бот!\n\n"
        "С чем я могу помочь:\n"
        "✔️ Забронировать место в автобусе.\n"
        "✔️ Посмотреть твою текущую бронь.\n"
        "✔️ Отменить запись.\n"
        "✔️ Предоставить информацию о маршрутах и способах доехать до ProductCamp.\n\n"
        "Чтобы начать, выберите нужную кнопку или введите команду /start."
    )
    keyboard = [
        [InlineKeyboardButton("Записаться на автобус", callback_data="book_bus")],
        [InlineKeyboardButton("Посмотреть свою бронь", callback_data="view_booking")],
        [InlineKeyboardButton("Отменить запись", callback_data="cancel_booking")],
        [InlineKeyboardButton("Как добраться?", callback_data="how_to_get_there")],
        [InlineKeyboardButton("FAQ", callback_data="render_faq")],
    ]
    
    # Добавляем кнопку выгрузки только для администраторов
    if username:
        try:
            # Получаем запись пользователя
            cell = passengers_sheet.find(username)
            headers = passengers_sheet.row_values(1)
            if "Role" in headers:
                role_col_idx = headers.index("Role") + 1
                role = passengers_sheet.cell(cell.row, role_col_idx).value
                if role and role.lower() == "admin":
                    keyboard.append([InlineKeyboardButton("Выгрузить данные", callback_data="export_buses")])
        except:
            pass

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.message:
        await update.message.reply_text(welcome_message, reply_markup=reply_markup)
    else:
        await update.callback_query.message.reply_text(welcome_message, reply_markup=reply_markup)

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == "book_bus":
        await step_select_direction(query, context)
    elif query.data == "view_booking":
        await view_booking(update, context)
    elif query.data == "cancel_booking":
        await cancel_booking(update, context)
    elif query.data.startswith("select_direction_"):
        direction = query.data.split("_", 2)[2]
        await show_buses_for_direction(query, context, direction)
    elif query.data.startswith("select_bus_"):
        bus_id = int(query.data.split("_")[2])
        await confirm_booking(update, context, bus_id)
    elif query.data.startswith("cancel_reservation_"):
        reservation_id = int(query.data.split("_")[2])
        await delete_reservation(update, context, reservation_id)
    elif query.data == "how_to_get_there":
        await show_how_to_get_there(update, context)
    elif query.data == "route_to_hotel":
        await show_how_route_to_hotel(update, context)
    elif query.data == "join_waiting_list":
        await add_user_to_waiting_list_callback(update, context)
    elif query.data == "waiting_list_back":
        await start(update, context)
    elif query.data.startswith("set_waiting_bus_"):
        await handle_select_waiting_bus(update, context)
    elif query.data.startswith("render_faq"):
        await render_faq(update, context)
    elif query.data == "export_buses":
        await export_buses(update, context)
    elif query.data == "back_to_menu":
        await start(update, context)

        
async def export_buses(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    import openpyxl
    """Асинхронная выгрузка данных в Excel с обработкой всех ошибок"""
    try:
        # Проверка прав доступа (теперь по роли admin)
        user = update.effective_user
        if not user:
            if update.callback_query:
                await update.callback_query.answer("⚠️ Доступ запрещен", show_alert=True)
            return
            
        # Получаем данные о пользователе из таблицы Passengers
        try:
            passengers = await asyncio.to_thread(passengers_sheet.get_all_records)
            user_record = next((p for p in passengers if p.get("Telegram_username", "") == user.username), None)
            if not user_record or user_record.get("Role", "").lower() != "admin":
                if update.callback_query:
                    await update.callback_query.answer("⚠️ Доступ запрещен", show_alert=True)
                return
        except Exception as e:
            raise Exception(f"Ошибка проверки прав доступа: {str(e)}")

        # Уведомление о начале процесса
        if update.callback_query:
            await update.callback_query.answer()
            msg = await update.callback_query.edit_message_text("🔄 Подготовка отчета...")
        else:
            msg = await update.message.reply_text("🔄 Подготовка отчета...")

        # Асинхронное получение данных
        try:
            buses, reservations, passengers, bus_owners = await asyncio.gather(
                asyncio.to_thread(buses_sheet.get_all_records),
                asyncio.to_thread(reservations_sheet.get_all_records),
                asyncio.to_thread(passengers_sheet.get_all_records),
                asyncio.to_thread(spreadsheet.worksheet("BusOwners").get_all_records)
            )
        except Exception as e:
            raise Exception(f"Ошибка получения данных из Google Sheets: {str(e)}")

        # Создаем временный файл
        temp_file = "temp_export.xlsx"
        
        try:
            # Создаем Excel-файл
            def generate_excel():
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # Удаляем дефолтный лист

                # Группируем пассажиров по автобусам
                bus_passengers = {}
                for res in reservations:
                    bus_id = str(res["Bus"])
                    passenger = next((p for p in passengers if str(p["ID"]) == str(res["Passenger"])), None)
                    if passenger:
                        if bus_id not in bus_passengers:
                            bus_passengers[bus_id] = []
                        bus_passengers[bus_id].append(passenger)

                # Создаем листы для каждого автобуса
                for bus in buses:
                    bus_id = str(bus["ID"])
                    sheet_name = f"Автобус {bus['Number']}"[:31]  # Ограничение длины имени листа
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Добавляем заголовок с информацией об автобусе
                    bus_info = [
                        f"Автобус: {bus.get('Number', '')}",
                        f"Маршрут: {bus.get('Departure_Place', '')} - {bus.get('Destination', '')}",
                        f"Направление: {bus.get('Direction'), ''}",
                        f"Дата/время: {bus.get('DepartureDate', '')} {bus.get('DepartureTime', '')}",
                        f"Вместимость: {bus.get('Capacity', '')}",
                    ]
                    
                    # Находим ответственных за автобус
                    owners = [o for o in bus_owners if str(o.get("BusID")) == bus_id]
                    responsible_persons = []
                    for owner in owners:
                        chief = next((p for p in passengers if str(p["ID"]) == str(owner.get("ChiefID"))), None)
                        if chief:
                            responsible_persons.append(
                                f"{chief.get('ФИО', '')} (@{chief.get('Telegram_username', '')})"
                            )
                    
                    if responsible_persons:
                        bus_info.append(f"Ответственные: {', '.join(responsible_persons)}")
                    
                    # Записываем информацию об автобусе
                    for line in bus_info:
                        ws.append([line])
                    
                    # Пустая строка для разделения
                    ws.append([])
                    
                    # Заголовки таблицы пассажиров
                    ws.append(["№", "ФИО", "Username", "Телефон", "Комментарий"])
                    
                    # Данные пассажиров
                    for idx, passenger in enumerate(bus_passengers.get(bus_id, []), start=1):
                        ws.append([
                            idx,
                            passenger.get("ФИО", ""),
                            f"@{passenger.get('Telegram_username', '')}",
                            passenger.get("Phone", ""),
                            passenger.get("Comment", "")
                        ])

                wb.save(temp_file)

            await asyncio.to_thread(generate_excel)
            await msg.edit_text("📤 Отправляем файл...")

            # Отправка файла
            with open(temp_file, "rb") as file:
                await context.bot.send_document(
                    chat_id=update.effective_chat.id,
                    document=file,
                    filename=f"bus_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    caption="✅ Отчет по автобусам"
                )
            
            await msg.delete()

        except Exception as e:
            error_msg = f"❌ Ошибка при генерации отчета: {str(e)}"
            await msg.edit_text(error_msg)
            return
        finally:
            # Удаляем временный файл
            if os.path.exists(temp_file):
                os.remove(temp_file)

    except Exception as e:
        error_msg = f"❌ Ошибка при выгрузке: {str(e)}"
        if update.callback_query:
            await update.callback_query.edit_message_text(error_msg)
        else:
            await update.message.reply_text(error_msg)


async def step_select_direction(query, context):
    buses = buses_sheet.get_all_records()

    # Получение уникальных направлений, которые есть у автобусов
    directions = sorted(set(b['Direction'] for b in buses if 'Direction' in b and b['Direction'].strip()))
    
    # Проверка, есть ли вообще направления
    if not directions:
        await query.edit_message_text("На данный момент нет доступных направлений.")
        return

    # Формируем кнопки с направлениями
    keyboard = [
        [InlineKeyboardButton(direction, callback_data=f"select_direction_{direction}")]
        for direction in reversed(directions)
    ]
    keyboard.append([InlineKeyboardButton("Назад", callback_data="back_to_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text("Выберите направление:", reply_markup=reply_markup)

async def show_buses_for_direction(query, context, direction):
    buses = buses_sheet.get_all_records()
    reservations = reservations_sheet.get_all_records()

    username = query.from_user.username
    passengers = passengers_sheet.get_all_records()
    passenger = next((p for p in passengers if p["Telegram_username"] == username), None)
    if not passenger:
        await query.edit_message_text(
            "К сожалению, я не вижу тебя в списках участиников кэмпа, "
            "для решение данного вопроса, обратись к своему старшему, либо напиши: @maximovd"
        )
        return

    # Проверка, есть ли уже регистрация у пользователя на данное направление
    existing_res = [
        r for r in reservations
        if r["Passenger"] == passenger["ID"] and r.get("Direction", "") == direction
    ]
    if existing_res:
        await query.edit_message_text("Вы уже зарегистрированы на это направление.")
        return

    # Фильтрация автобусов по выбранному направлению
    buses_for_direction = [b for b in buses if b.get("Direction") == direction]

    # Собираем информацию о занятости и статусе автобусов
    bus_info_list = []
    for bus in buses_for_direction:
        bus_id = bus["ID"]
        capacity_value = bus.get("Capacity", "")
        try:
            capacity = int(capacity_value.strip())
        except:
            capacity = 0
        bus_reservations = [r for r in reservations if r["Bus"] == bus_id]
        booked = len(bus_reservations)
        free_places = capacity - booked
        bus['FreePlaces'] = free_places
        has_places = free_places > 0
        bus_info_list.append((bus, has_places))

    # Отобрать все автобусы даже если они заполнены
    all_buses = buses_for_direction

    # Формируем сообщение
    message = f"Все автобусы для направления {direction}:\n\n"

    for bus in all_buses:
        capacity_str = bus.get("Capacity", "0")
        try:
            capacity = int(capacity_str)
        except:
            capacity = 0
        booked = sum(1 for r in reservations if r["Bus"] == bus["ID"])
        free = capacity - booked
        if free <= 0:
            status = "Свободных мест: 0"
        else:
            status = f"Свободных мест: {free}"
        message += (
            f"Автобус {bus['ID']} ({bus['DepartureDate']} {bus['DepartureTime']}): {status}\n"
        )

    # Создавать кнопки для всех автобусов:
    # - тех, у которых есть свободные места — кнопка для регистрации
    # - тех, что заполнены — кнопка для постановки в очередь
    keyboard = []

    for bus in all_buses:
        capacity_str = bus.get("Capacity", "0")
        try:
            capacity = int(capacity_str)
        except:
            capacity = 0
        booked = sum(1 for r in reservations if r["Bus"] == bus["ID"])
        free = capacity - booked
        if free > 0:
            # Есть свободные места — кнопка "Записаться"
            keyboard.append([
                InlineKeyboardButton(
                    f"Автобус {bus['Number']} - {free} мест",
                    callback_data=f"select_bus_{bus['ID']}"
                )
            ])
        else:
            # Полностью заполнен — кнопка "Записан и в очередь"
            keyboard.append([
                InlineKeyboardButton(
                    f"Автобус {bus['Number']} - в лист ожидания",
                    callback_data=f"set_waiting_bus_{bus['ID']}"
                )
            ])

    keyboard.append([InlineKeyboardButton("Назад", callback_data="back_to_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(message, reply_markup=reply_markup)


async def view_booking(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    username = query.from_user.username
    passengers = passengers_sheet.get_all_records()
    passenger = next((p for p in passengers if p["Telegram_username"] == username), None)
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if not passenger:
        await query.edit_message_text(
            "К сожалению, я не вижу тебя в списках участиников кэмпа, "
            "для решение данного вопроса, обратись к своему старшему, либо напиши: @maximovd",
            reply_markup=reply_markup,
        )
        return

    reservations = reservations_sheet.get_all_records()
    buses = buses_sheet.get_all_records()

    user_res = [r for r in reservations if r["Passenger"] == passenger["ID"]]
    if not user_res:
        await query.edit_message_text("Вы не записаны ни на один автобус", reply_markup=reply_markup)
        return

    message = "Ваши записи:\n\n"
    keyboard = []
    for res in user_res:
        bus = next((b for b in buses if b["ID"] == res["Bus"]), None)
        if bus:
            message += (
                f"Автобус: {bus['Number']} ({bus['DepartureDate']} {bus['DepartureTime']}) "
                f"({bus['Departure_Place']}-{bus['Destination']})-{bus['Direction']}\n"
            )
            keyboard.append([InlineKeyboardButton(f"Отменить бронь: Автобус: {bus['Number']}-{bus['Direction']}", callback_data=f"cancel_reservation_{res['ID']}")])
    
    keyboard.append([InlineKeyboardButton("Назад", callback_data="back_to_menu")])

    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(message, reply_markup=reply_markup)

async def cancel_booking(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    username = query.from_user.username
    passengers = passengers_sheet.get_all_records()
    passenger = next((p for p in passengers if p["Telegram_username"] == username), None)
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if not passenger:
        await query.edit_message_text(
            "К сожалению, я не вижу тебя в списках участиников кэмпа, "
            "для решение данного вопроса, обратись к своему старшему, либо напиши: @maximovd",
            reply_markup=reply_markup,
        )
        return

    reservations = reservations_sheet.get_all_records()
    buses = buses_sheet.get_all_records()

    user_res = [r for r in reservations if r["Passenger"] == passenger["ID"]]
    if not user_res:
        await query.edit_message_text("У вас нет записей для отмены", reply_markup=reply_markup)
        return

    keyboard = []
    for res in user_res:
        bus = next((b for b in buses if b["ID"] == res["Bus"]), None)
        if bus:
            button_text = (
                f"Автобус {bus['Number']} ({bus['DepartureDate']} {bus['DepartureTime']}) Направление: {bus['Direction']}"
            )
            keyboard.append(
                [InlineKeyboardButton(
                    button_text,
                    callback_data=f"cancel_reservation_{res['ID']}"
                )]
            )

    keyboard.append([InlineKeyboardButton("Назад", callback_data="back_to_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(
        "Выберите запись для отмены:", reply_markup=reply_markup
    )

async def delete_reservation(update, context, reservation_id):
    """Удаление записи"""
    query = update.callback_query
    username = query.from_user.username
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    reservations = reservations_sheet.get_all_records()
    reservation = next((r for r in reservations if str(r["ID"]) == str(reservation_id)), None)

    if not reservation:
        await query.edit_message_text("Запись не найдена.", reply_markup=reply_markup)
        return

    passengers = passengers_sheet.get_all_records()
    passenger = next((p for p in passengers if p["Telegram_username"] == username), None)

    if not passenger or str(reservation["Passenger"]) != str(passenger["ID"]):
        await query.edit_message_text("Ошибка: эта запись вам не принадлежит.", reply_markup=reply_markup)
        return

    cell = reservations_sheet.find(str(reservation_id))
    reservations_sheet.delete_rows(cell.row)
    await query.edit_message_text("Ваша запись отменена.", reply_markup=reply_markup)

async def show_how_to_get_there(update: Update, context: ContextTypes.DEFAULT_TYPE):
    info_message = (
        "Планируете добираться до мероприятия самостоятельно?\n"
        "Мы собрали для вас удобные варианты проезда различными видами транспорта – выбирайте свой!\n"
        "**Адрес отеля**: Московская область, Ногинский район, д. Новая Купавна, местечко Родинки, ул. Сиреневая, д.21 стр.1\n"
        "**Метки на картах**: [Гугл-карта](https://goo.gl/maps/7doSnWnGg4mb8QqG6), [Яндекс-карта](https://yandex.ru/maps/-/CCUWeOGxwD)\n\n"
        "На электричке (время в пути ~2,5 часа)\n"
        "от Курского или Белорусского вокзала до станции Купавна (цена билета ~150 руб), \n"
        "затем 20-30 мин на такси (~650 руб.) до отеля Ареал\n"
        "или автобусе 37щ до ост. Улица Орлова, 26 и пешком/на такси (~350 руб.) до отеля 3 км.\n\n"
        "**Расписание электричек**: [туда](https://ticket.rzd.ru/searchresults/v/1/5a323c29340c7441a0a556bb/5cd18d837081a600437b02f2/) и [обратно](https://ticket.rzd.ru/searchresults/v/1/5cd18d837081a600437b02f2/5a323c29340c7441a0a556bb/) пожалуйста, внимательно проверяйте станцию отправления и прибытия в Москве, расписание указано для двух вокзалов.\n\n"
        "На автобусе (время в пути ~1,5 часа)\n"
        "от ст. м. Партизанская/МЦК Измайлово (1 выход из метро) - автобусы №322 или №399 или №444\n"
        "от м. Новогиреево маршрутное такси №1209к, №587к или №886к (6 выход из метро); \n"
        "до ост. Новая Купавна, перейти Горьковское шоссе, \n"
        "далее на такси 5-7 мин (~250 руб.) или пешком около 2,3 км по указателям (для спортивных участников).\n\n"
        "**Расписание автобусов**: [туда](https://rasp.yandex.ru/search/bus/?fromId=c213&fromName=%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B0&toId=c33762&toName=%D0%9D%D0%BE%D0%B2%D0%B0%D1%8F+%D0%9A%D1%83%D0%BF%D0%B0%D0%B2%D0%BD%D0%B0&when=23+%D0%BC%D0%B0%D1%8F) и [обратно](https://rasp.yandex.ru/search/bus/?fromId=c33762&fromName=%D0%9D%D0%BE%D0%B2%D0%B0%D1%8F+%D0%9A%D1%83%D0%BF%D0%B0%D0%B2%D0%BD%D0%B0&toId=c213&toName=%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B0&when=25+%D0%BC%D0%B0%D1%8F) пожалуйста, внимательно проверяйте даты отправления и прибытия, данные актуальны для пятницы 23 мая и воскресенья 25 мая соответственно.\n\n"
        "На машине (время в пути ~1 час)\n"
        "на территории отеля есть бесплатная парковка на 190 мест, находится перед въездом в отель (пропуск не требуется). Подъезд со стороны Горьковского шоссе. \n"
        "Если есть свободные места и вы можете взять попутчиков, укажите это в чате: https://t.me/c/123456789/101112, начните знакомство еще до кэмпа!😉\n\n"
        "**Из аэропортов Москвы**\n"
        "Шереметьево (~2,5 часа)\n"
        "[Аэроэкспрессе](https://aeroexpress.ru/)(~650 руб.) до Окружной, пересесть на МЦК до ст. Измайловская;\n"
        "или [Экспресс автобус](https://aeroexpress.ru/) (~400 руб.) до Ховрино, затем на МЦК до ст. Партизанская.\n"
        "далее на автобусе №322, №399 или №444 до ост. Новая Купавна, далее на такси (~250 руб.) или пешком около 2,3 км.\n"
        "Внуково (~2,5 часа)\n"
        "на метро от ст. Аэропорт Внуково до ст. Партизанская, далее на автобусе №322, №399 или №444 до ост. Новая Купавна, далее на такси (~250 руб.) или пешком 2,3 км.\n"
        "Домодедово (~2,5 часа)\n"
        "на [Аэроэкспрессе](https://aeroexpress.ru/) (~650 руб.) до Верхних Котлов, пересесть на МЦК до ст. Измайловская, далее на автобусе (№322, 399 или №444) до ост. Новая Купавна, затем на такси (~250 руб.) или пешком 2,3 км.\n"
        "❗️ Учитывайте, пожалуйста, что в пятницу вечером и субботу утром на данном направлении могут быть пробки из Москвы, а в воскресенье вечером - в Москву, что может вызвать задержки в пути❗️"
    )
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(info_message, reply_markup=reply_markup, parse_mode="Markdown")

async def render_faq(update: Update, context: ContextTypes.DEFAULT_TYPE):
    faq_text = (
        "1. **Как записаться?** Открой бота → нажми /start → записаться на автобус → выбери направление → выбери автобус → нажми «Записаться». Готово!\n"
        "\n"
        "2. **Что, если я хочу и туда, и обратно?** Нужно пройти процесс записи дважды — сначала в одном направлении, потом в другом.\n"
        "\n"
        "3. **Что, если в нужном автобусе нет мест?** Это значит, что все места на выбранный автобус уже заняты. Однако ты всё равно можешь выбрать его в боте, чтобы попасть в лист ожидания. Если кто-то откажется от своей брони, ты получишь уведомление от бота и сможешь забронировать освободившееся место.\n"
        "\n"
        "4. **Как посмотреть свою бронь?** Через кнопку «Мои записи» в меню. Там же можно отменить любую бронь.\n"
        "\n"
        "5. **Могу ли я изменить автобус?** Да — сначала отмени текущую запись, потом выбери другой автобус.\n"
        "\n"
        "6. **Почему бот не пускает?** Вероятно, твой Telegram-ник не совпадает с базой волонтёров. Обратись к своему старшему или проверь, с какого аккаунта ты зашёл.\n"
        "\n"
        "7. **Куда писать, если ничего не работает?** Если бот не пускает или что-то сломалось, напиши Данилу @maximovd или в [общий чат](https://t.me/c/2339287372/1173) ProductCamp во вкладку «Трансфер»\n"
        "\n"
        "8. **Что значит «в листе ожидания»?** Это значит, что автобус, на который ты хотел(а) записаться, уже заполнен. Однако бот всё равно добавил тебя в лист ожидания. Если кто-то отменит бронь — ты получишь уведомление и сможешь забронировать освободившееся место.\n"
        "\n"
        "9. **Можно ли записаться на несколько автобусов сразу?** Нет, по одному направлению можно иметь только одну активную запись. Сначала отмени предыдущую, если хочешь сменить автобус.\n"
        "\n"
        "10.**А если я хочу добраться до ProductCamp самостоятельно?** В боте ты сможешь найти информацию о маршрутах и способах доехать до ProductCamp. Для этого просто зайди в бота, нажми «/start», а затем в меню нажми на кнопку «Как добраться до Product Camp».\n"
        "\n"
        "11.**Почему автобусов так мало и расписание такое неудобное?** Мы понимаем, что расписание может подойти не всем. У нас ограниченный бюджет, и мы старались распределить автобусы так, чтобы охватить максимальное число кэмпчан. Надеемся, что выбранные будут удобными для большинства.\n"
        "\n"
        "Пользуйся ботом, экономь своё время и не забудь забронировать место заранее! Если будут вопросы — мы будем в общем чате ProductCamp во вкладке «Трансфер» 😊\n"
    )
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(faq_text, parse_mode="Markdown", reply_markup=reply_markup)

async def show_how_route_to_hotel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    image_url = "https://lh7-rt.googleusercontent.com/docsz/AD_4nXeTgIHj8xL5841KSiK2TweN4zWuVuVyl9kEzjhLycPcNMlVY3Q2K4ArRzdy1ZpGIXbS6-hYWmNEpYq4h6B2py8EmfDX1w3K225docCZAXI3Esh6iKHBPKhad-QUSq1ND68n4HhE0w?key=pfElDP_kPhatX9dfoNbfQj_I"

    info_text = (
    "Мероприятие в конгресс-отеле [Ареал](https://www.areal-hotel.ru/about-the-hotel/) в Московской области.\n"
    "Адрес: Сиреневая ул., 21, микрорайон Родинки, д. Новая Купавна.\n"
    "Метки на картах: [гугл-карта](https://goo.gl/maps/7doSnWnGg4mb8QqG6), [яндекс-карта](https://yandex.ru/maps/-/CCUWeOGxwD)\n\n"
    "🚌 **Способы добраться до Отеля:**\n"
    "На организованном трансфере, время в пути около 1 часа: в соответствии с расписанием трансфера будут организованы автобусы от станции метро Партизанская.\n"
    "Ссылка на точку: https://clck.ru/348UK5"
    )
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_media(
        media=InputMediaPhoto(media=image_url, caption=info_text, parse_mode="Markdown"),
        reply_markup=reply_markup,
    )


async def confirm_booking(update, context, bus_id):
    """Подтверждение записи на автобус с проверкой занятости мест"""
    query = update.callback_query
    username = query.from_user.username
    passengers_sheet = spreadsheet.worksheet("Passengers")
    buses_sheet = spreadsheet.worksheet("Buses")
    reservations_sheet = spreadsheet.worksheet("Reservations")
    wating_list_sheet = spreadsheet.worksheet("WaitingList")
    
    passengers = passengers_sheet.get_all_records()
    passenger = next((p for p in passengers if p["Telegram_username"] == username), None)
    query = update.callback_query
    keyboard = [
        [InlineKeyboardButton("Назад", callback_data="back_to_menu")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if not passenger:
        await query.edit_message_text(
            "К сожалению, я не вижу тебя в списках участиников кэмпа, "
            "для решение данного вопроса, обратись к своему старшему, либо напиши: @maximovd",
            reply_markup=reply_markup,
        )
        return

    buses = buses_sheet.get_all_records()
    reservations = reservations_sheet.get_all_records()

    bus = next((b for b in buses if b["ID"] == bus_id), None)
    if not bus:
        await query.edit_message_text("Автобус не найден.")
        return

    # Получаем вместимость автобуса
    capacity_value = bus.get("Capacity", "")
    if isinstance(capacity_value, str):
        capacity_str = capacity_value.strip()
        try:
            capacity = int(capacity_str)
        except ValueError:
            capacity = 0
    elif isinstance(capacity_value, int):
        capacity = capacity_value
    else:
        capacity = 0

    # Проверка, заняты ли все места
    bus_reservations = [r for r in reservations if int(r["Bus"]) == int(bus_id)]
    if len(bus_reservations) >= capacity:
        await query.edit_message_text("Все места в автобусе уже заняты.", reply_markup=reply_markup)
        return

    # Проверка, есть ли регистрация на этот автобус и направление
    direction_value = bus.get("Direction", "")
    existing_res = [
        r for r in reservations
        if r["Passenger"] == str(passenger["ID"]) and r["Bus"] == str(bus_id) and r.get("Direction", "") == direction_value
    ]
    if existing_res:
        await query.edit_message_text("Вы уже зарегистрированы на этот автобус и направление.")
        return

    # Создание новой брони
    existing_ids = [int(val) for val in reservations_sheet.col_values(1) if val.replace('.', '', 1).isdigit()]
    max_id = max(existing_ids) if existing_ids else 0
    new_id = str(max_id + 1)
    direction = bus.get("Direction", "")

    new_reservation = [
        new_id,
        str(passenger["ID"]),
        str(bus_id),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        direction,
    ]
    reservations_sheet.append_row(new_reservation)
    await query.edit_message_text(
        f"Вы успешно записаны на автобус: {bus['Number']} "
        f"({bus['DepartureDate']} {bus['DepartureTime']}) "
        f"{bus['Departure_Place']}-{bus['Destination']}",
        reply_markup=reply_markup,
    )

    waiting_records = wating_list_sheet.get_all_records()

    for wait in waiting_records:
        if (str(wait.get("PassengerID")) == str(passenger["ID"]) and 
            str(wait.get("BusID")) == str(bus_id) and 
            wait.get("Status") == "Waiting"):
            
            # Находим и обновляем запись
            try:
                cell = wating_list_sheet.find(str(wait["ID"]))
                status_col = wating_list_sheet.find("Status").col
                wating_list_sheet.update_cell(cell.row, status_col, "Confirmed")
                
                # Также обновляем NotificationSent, чтобы не отправлять повторные уведомления
                notification_col = wating_list_sheet.find("NotificationSent").col
                wating_list_sheet.update_cell(cell.row, notification_col, "Yes")
                
                logging.info(f"Обновлена запись в WaitingList: ID {wait['ID']} изменен на Confirmed")
            except Exception as e:
                logging.error(f"Ошибка при обновлении WaitingList: {str(e)}")
            return


async def process_waiting_list(application: Application, single_notification: bool = True):
    """Автоматическая проверка и оповещение о свободных местах с учетом статусов.
    Если single_notification=True, отправляет уведомление только самому раннему ожидающему."""
    try:
        # Получаем все необходимые данные
        ws = spreadsheet.worksheet("WaitingList")
        reservations_ws = reservations_sheet
        buses = await asyncio.to_thread(buses_sheet.get_all_records)
        reservations = await asyncio.to_thread(reservations_ws.get_all_records)
        waiting_records = await asyncio.to_thread(ws.get_all_records)
        passengers = await asyncio.to_thread(passengers_sheet.get_all_records)

        # Подготавливаем данные о вместимости автобусов
        bus_capacity = {}
        for b in buses:
            try:
                bus_capacity[b['ID']] = int(b.get('Capacity', '0'))
            except (ValueError, TypeError):
                bus_capacity[b['ID']] = 0

        # Подсчитываем количество броней для каждого автобуса
        bus_reserved_counts = {}
        for b in buses:
            bus_reserved_counts[b['ID']] = sum(1 for r in reservations if str(r["Bus"]) == str(b['ID']))

        # Текущее время для проверки 2-х суток
        current_time = datetime.now()

        # Если нужно отправить только одно уведомление самому раннему ожидающему
        if single_notification:
            eligible_records = []
            
            for wait in waiting_records:
                if (wait.get("Status") == "Waiting" and 
                    wait.get("NotificationSent") == "No" and 
                    wait.get("RequestTime")):
                    
                    try:
                        request_time = datetime.strptime(wait["RequestTime"], "%Y-%m-%d %H:%M:%S")
                        eligible_records.append((request_time, wait))
                    except (ValueError, TypeError):
                        continue
            
            if eligible_records:
                # Находим запись с самым ранним временем
                eligible_records.sort(key=lambda x: x[0])
                earliest_record = eligible_records[0][1]
                waiting_records = [earliest_record]  # Обрабатываем только эту запись
            else:
                return  # Нет подходящих записей для уведомления

        for wait in waiting_records:
            # Пропускаем записи не в статусе Waiting (для общего случая)
            if wait.get("Status") != "Waiting":
                continue

            bus_id = wait.get("BusID")
            passenger_id = wait.get("PassengerID")
            
            # Проверка на существование автобуса
            if bus_id not in bus_capacity:
                continue

            # Проверка времени для сброса NotificationSent
            request_time_str = wait.get("RequestTime")
            if request_time_str:
                try:
                    request_time = datetime.strptime(request_time_str, "%Y-%m-%d %H:%M:%S")
                    time_diff = current_time - request_time
                    
                    # Если прошло более 2 суток и статус Waiting
                    if time_diff.total_seconds() > 172800 and wait.get("NotificationSent") == "Yes":
                        # Находим и обновляем запись
                        cell = ws.find(str(wait["ID"]))
                        ws.update_cell(cell.row, ws.find("NotificationSent").col, "No")
                        continue
                except (ValueError, TypeError):
                    pass

            # Пропускаем если уведомление уже отправлено (для общего случая)
            if wait.get("NotificationSent") == "Yes" and not single_notification:
                continue

            # Проверяем наличие свободных мест
            free_places = bus_capacity[bus_id] - bus_reserved_counts.get(bus_id, 0)
            if free_places <= 0:
                continue

            # Находим пассажира и автобус
            passenger = next((p for p in passengers if str(p["ID"]) == str(passenger_id)), None)
            bus = next((b for b in buses if str(b["ID"]) == str(bus_id)), None)

            if passenger and bus and passenger.get("ChatID"):
                # Создаем клавиатуру для подтверждения
                keyboard = [
                    [InlineKeyboardButton(
                        "Подтвердить бронь",
                        callback_data=f"select_bus_{bus_id}"
                    )]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                # Отправляем уведомление
                try:
                    await application.bot.send_message(
                        chat_id=passenger["ChatID"],
                        text=f"🚌 Место на автобус {bus['Number']} ({bus['DepartureDate']} {bus['DepartureTime']}) теперь доступно!\n"
                             "❗У вас есть 10 минут, чтобы подтвердить бронь, после бот отправит пуш следующему в листе ожидания. \n"
                             "Нажмите кнопку ниже чтобы подтвердить бронь:",
                        reply_markup=reply_markup
                    )
                    
                    # Обновляем статус NotificationSent
                    cell = ws.find(str(wait["ID"]))
                    notification_col = ws.find("NotificationSent").col
                    ws.update_cell(cell.row, notification_col, "Yes")
                    
                    # Логируем успешную отправку
                    logging.info(f"Уведомление отправлено пассажиру {passenger_id} для автобуса {bus_id}")
                        
                except Exception as e:
                    logging.error(f"Ошибка отправки уведомления пассажиру {passenger_id}: {str(e)}")

    except Exception as e:
        logging.error(f"Ошибка в process_waiting_list: {str(e)}")

async def post_init(application: Application):
    asyncio.create_task(periodic(application))


async def confirm_waiting(update, context):
    user = update.message.from_user
    passenger = next((p for p in passengers_sheet.get_all_records() if p["Telegram_username"] == user.username), None)
    if not passenger:
        await update.message.reply_text("Пользователь не найден.")
        return
    ws = spreadsheet.worksheet("WaitingList")
    all_records = ws.get_all_records()
    for rec in all_records:
        if rec["PassengerID"] == str(passenger["ID"]) and rec["Status"] == "Confirmed":
            await update.message.reply_text("Вы уже подтвердили бронь.")
            return
        elif rec["PassengerID"] == str(passenger["ID"]) and rec["Status"] == "Waiting":
            ws.update_cell(ws.find(rec["ID"]).row, 5, "Confirmed")
            # Создать бронь
            reservations = reservations_sheet.get_all_records()
            buses = buses_sheet.get_all_records()
            bus_id = rec["BusID"]
            bus = next((b for b in buses if b["ID"] == bus_id), None)
            if not bus:
                await update.message.reply_text("Автобус не найден.")
                return
            current_res = [r for r in reservations if r["Bus"] == bus_id and r.get("Status") == "Booked"]
            if len(current_res) >= int(bus.get("Capacity", "0").strip()):
                await update.message.reply_text("Места в автобусе уже заняты.")
                return
            existing_ids = [int(r["ID"]) for r in reservations if r["ID"].isdigit()]
            new_id = str(max(existing_ids) + 1 if existing_ids else 1)
            reservations_sheet.append_row([
                new_id,
                str(passenger["ID"]),
                str(bus_id),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Booked",
                bus.get("Direction", "")
            ])
            await update.message.reply_text("Бронь подтверждена. Удачи!")
            return
    await update.message.reply_text("Нет ожидающих для подтверждения или уже подтверждены.")

async def add_user_to_waiting_list_callback(update, context):
    query = update.callback_query
    user = query.from_user
    passenger = next((p for p in passengers_sheet.get_all_records() if p["Telegram_username"] == user.username), None)
    if not passenger:
        await query.answer("Пользователь не найден.")
        return
    # Предложить выбрать автобус для постановки в очередь
    buses = buses_sheet.get_all_records()
    if not buses:
        await query.answer("Нет автобусов для очереди.")
        return
    keyboard = [
        [InlineKeyboardButton(f"Автобус {b['Number']} ({b['DepartureDate']} {b['DepartureTime']})", callback_data=f"set_waiting_bus_{b['ID']}")]
        for b in buses
    ]
    keyboard.append([InlineKeyboardButton("Отмена", callback_data="back_to_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text("Выберите автобус для постановки в очередь:", reply_markup=reply_markup)

async def handle_select_waiting_bus(update, context):
    query = update.callback_query

    if query.data.startswith("set_waiting_bus_"):
        bus_id = query.data.split("_", 3)[3]
        user = query.from_user
        passenger = next((p for p in passengers_sheet.get_all_records() if p["Telegram_username"] == user.username), None)

        if not passenger:
            await query.answer("Пользователь не найден.")
            return

        ws = spreadsheet.worksheet("WaitingList")
        existing = ws.findall(str(passenger["ID"]))

        for cell in existing:
            row = cell.row
            b_id_cell = ws.cell(row, 3)
            status_cell = ws.cell(row, 5)
            if b_id_cell.value == bus_id and status_cell.value == "Waiting":
                await query.edit_message_text("Вы уже в очереди на этот автобус.")
                await query.answer()
                return

        existing_records = ws.get_all_records()
        print(existing_records)
        new_id = max([int(r["ID"]) for r in existing_records]+[0]) + 1
        ws.append_row([
            str(new_id),
            str(passenger["ID"]),
            str(bus_id),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Waiting",
            "No"
        ])
        await query.edit_message_text("Вы поставлены в очередь на выбранный автобус. Ожидайте уведомлений о доступных местах.")
        await query.answer()

async def handle_waiting_list_entry(update, context):
    # Пользователь вызвал команду или кнопку для вхождения в лист ожидания
    await add_user_to_waiting_list_callback(update, context)

async def handle_waiting_list_back(update, context):
    await start(update, context)

async def periodic(application):
    while True:
        try:
            await process_waiting_list(application)
        except Exception as e:
            print(f"Ошибка: {e}")
        await asyncio.sleep(600)  # TODO: Вынести это в переменную окружения


def main():
    application = Application.builder().token(TOKEN).post_init(post_init).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(CommandHandler("confirm", confirm_waiting))
    application.add_handler(CommandHandler("wait", handle_waiting_list_entry))
    # Периодическая обработка листа ожидания


    application.run_polling()

if __name__ == "__main__":
    main()
